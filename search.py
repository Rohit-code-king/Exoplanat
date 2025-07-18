from tkinter import *
import openpyxl
from PIL import Image, ImageTk

search_root = None

def set_background(root, image_path):
    img = Image.open(image_path)
    window_width = root.winfo_screenwidth()
    window_height = root.winfo_screenheight()
    img = img.resize((window_width, window_height), Image.ANTIALIAS)
    img = ImageTk.PhotoImage(img)
    background_frame = Frame(root, width=window_width, height=window_height)
    background_frame.pack(side="bottom", fill="both", expand=True)
    background_label = Label(background_frame, image=img)
    background_label.image = img
    background_label.place(relwidth=1, relheight=1)
    background_frame.img = img

def create_search_frame(root):
    search_frame = Frame(root, bg="white", width=400, height=100)
    search_frame.place(relx=0.5, rely=0.1, anchor="n")
    Label(search_frame, text="Search Data", font=("freestyle script", 20)).pack()
    searched = StringVar()
    entry1 = Entry(search_frame, textvariable=searched, font=("verdana", 12), width=30)
    entry1.pack(pady=10)
    Button(search_frame, text="Submit", cursor='hand2', command=lambda: search_name(searched.get())).pack()

def search():
    global search_root
    search_root = Tk()
    search_root.title('Search')
    set_background(search_root, "bg.png")
    create_search_frame(search_root)
    search_root.mainloop()

def search_name(query):
    obj = openpyxl.load_workbook(r"C:\Users\rohit\project\exoplanet.xlsx")
    objj = obj.active
    data_found = False
    for i in range(8054):
        if objj.cell(row=i + 1, column=4).value == query:
            data = [
                objj.cell(row=i + 1, column=2).value,
                objj.cell(row=i + 1, column=3).value,
                objj.cell(row=i + 1, column=4).value,
                objj.cell(row=i + 1, column=5).value,
                objj.cell(row=i + 1, column=6).value,
                objj.cell(row=i + 1, column=7).value,
                objj.cell(row=i + 1, column=8).value,
                objj.cell(row=i + 1, column=9).value,
                objj.cell(row=i + 1, column=10).value,
                objj.cell(row=i + 1, column=11).value,
                objj.cell(row=i + 1, column=12).value,
                objj.cell(row=i + 1, column=13).value,
                objj.cell(row=i + 1, column=14).value
            ]
            show_search_results(data)
            data_found = True
            break
    if not data_found:
        not_found_message = Label(search_root, text="Data not found.", font=("freestyle script", 20), bg="white")
        not_found_message.place(relx=0.5, rely=0.5, anchor="center")  # Center the message
    else:
        # Data found, so we need to destroy the "Data not found" message if it's visible
        destroy_not_found_message()

def show_search_results(data):
    global search_root
    destroy_not_found_message()
    search_root.winfo_children()[0].destroy()
    set_background(search_root, "bg.png")
    labels = [
        "Kepler id", "Kepoi_name", "Kepler name", "Exo status",
        "Orbital Period", "Impact Parameter", "Transit duration",
        "Radii", "Orbit semi-major", "Temperature", "Star temp",
        "Star radii", "Star mass"
    ]
    results_frame = Frame(search_root, bg="white")
    results_frame.place(relx=0.5, rely=0.05, anchor="n")
    for label, value in zip(labels, data):
        label_widget = Label(results_frame, text=f"{label}: {value}", font=("freestyle script", 20))
        label_widget.pack()
        Label(results_frame, text="").pack()
    results_frame.update_idletasks()
    canvas_height = results_frame.winfo_height()
    canvas_height += len(labels) * 3
    results_frame.config(height=canvas_height)

def destroy_not_found_message():
    # Destroy the "Data not found" message if it exists
    for widget in search_root.winfo_children():
        if isinstance(widget, Label) and widget.cget("text") == "Data not found.":
            widget.destroy()

search()
